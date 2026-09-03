# /// script
# requires-python = ">=3.11"
# dependencies = ["openpyxl>=3.1"]
# ///
"""Compare the tool's answers with the human labels on the calibration set.

Usage:  uv run calibrate.py results/calibration.csv
Prints agreement per gate and for the overall decision under both decision rules, then every disagreement.
"""
import csv, sys
from pathlib import Path

HERE = Path(__file__).parent
LABELS = HERE / "data/calibration/reviewed_candidates_15.csv"

def main(results_csv: str) -> None:
    truth = {r["Synthetic ID"]: r for r in csv.DictReader(open(LABELS, encoding="utf-8-sig"))}
    pred = {r["Synthetic ID"]: r for r in csv.DictReader(open(results_csv, encoding="utf-8-sig"))}
    ids = [i for i in truth if i in pred]
    n = len(ids)

    def agree(col=None, f=None):
        f = f or (lambda r: r[col])
        return sum(f(truth[i]) == f(pred[i]) for i in ids)

    both = lambda r: "Progress" if r["Safety motivation"] == "Pass" and r["Overall impressiveness"] == "Pass" else "Do not progress"
    imp_only = lambda r: "Progress" if r["Overall impressiveness"] == "Pass" else "Do not progress"
    print(f"Calibration set: {n} labelled candidates\n")
    print(f"  Safety motivation        {agree('Safety motivation')}/{n}")
    print(f"  Overall impressiveness   {agree('Overall impressiveness')}/{n}")
    print(f"  Overall decision, rubric rule (both gates)         {agree(None, both)}/{n}")
    print(f"  Overall decision, tool rule (impressiveness gate)  {sum(imp_only(pred[i]) == truth[i]['Overall decision'] for i in ids)}/{n}")
    write_xlsx(ids, truth, pred, Path(results_csv).with_name("calibration_comparison.xlsx"))
    print("\nDisagreements (human -> tool):")
    for i in ids:
        t, p = truth[i], pred[i]
        for col in ["Safety motivation", "Overall impressiveness"]:
            if t[col] != p[col]:
                reason = p["Safety reason" if col.startswith("Safety") else "Impressiveness reason"]
                print(f"  {i} {t['Synthetic name']:<14} {col:<23} {t[col]} -> {p[col]}\n      tool: {reason}")

def write_xlsx(ids, truth, pred, path):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    cols = ["Synthetic ID", "Synthetic name", "Human: Safety", "Tool: Safety", "Human: Impressiveness", "Tool: Impressiveness",
            "Human: Decision", "Tool: Decision", "Tool: Priority", "Tool: Safety reason", "Tool: Impressiveness reason", "Tool: Flags"]
    wb = Workbook(); ws = wb.active; ws.title = "Human vs tool"
    ws.append(cols)
    for c in ws[1]:
        c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="D9E2F3")
    red = PatternFill("solid", fgColor="F8CBAD")
    for i in ids:
        t, p = truth[i], pred[i]
        ws.append([i, t["Synthetic name"], t["Safety motivation"], p["Safety motivation"], t["Overall impressiveness"],
                   p["Overall impressiveness"], t["Overall decision"], p["Overall decision"], p["Priority"],
                   p["Safety reason"], p["Impressiveness reason"], p["Flags"]])
        r = ws.max_row
        for hc, tc in [(3, 4), (5, 6), (7, 8)]:
            if ws.cell(r, hc).value != ws.cell(r, tc).value:
                ws.cell(r, tc).fill = red
    for i, w in enumerate([14, 16, 13, 13, 13, 13, 16, 16, 10, 55, 55, 18], 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "C2"
    wb.save(path)


if __name__ == "__main__":
    main(sys.argv[1] if len(sys.argv) > 1 else str(HERE / "results/calibration.csv"))
